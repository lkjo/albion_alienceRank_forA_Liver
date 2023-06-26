from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import time
import os
#目前差兩個部分
#1. 讀取xml or yml 設定檔 fin
#2. 讀取excel
    #a. 讀取A：帳號
    #b. 讀取F：目前積分
    #c. 讀取K,L,M,N,O：活動類型，參與該活動紀錄+1
#3. 寫入excel
    #a. 寫入F：目前積分
    #b. 讀取K,L,M,N,O：活動類型，參與該活動紀錄+1
def ipCheck(player,standardIpScore):#確認ip是否足夠 透過xml決定要不要checkip
    ipScore = player.find('div', class_='rs-table-cell-group rs-table-cell-group-scroll')
    ipScore = player.find('div', class_='rs-table-cell')
    ipScore = ipScore.text
    if ipScore >= standardIpScore:
        return True #可以拿積分
    return False #不能拿積分
def readXml():#讀取xml並將必須資料進行回傳
    tree = ET.parse('data.xml')
    root = tree.getroot()
    dataDict = {}
    for element in root:
        key, value = element.tag,element.text
        dataDict[key] = value
    return dataDict
def readExcel(nameSet = set(),score = 0):#讀取excel並寫入積分
    target = "AL人事.xlsx"
    wb = load_workbook(target)
    sheet = wb['積分表格']
    nameList = sheet.iter_rows()
    count = 1
    for cell in nameList:
        if cell[0].value == None:
            break
        print(sheet[cell[4].coordinate].value)
        if cell[0].value.upper() in nameSet:
            tempA,tempB = str(int(sheet[cell[4].coordinate].value) + score),str(int(sheet[cell[5].coordinate].value) + score)
            sheet[cell[4].coordinate] = tempA
            sheet[cell[5].coordinate] = tempB
            nameSet.remove(cell[0].value.upper())
        count += 1
    nameSet = list(nameSet)
    while nameSet:
        sheet.cell(row = count,column=1, value=nameSet[0])
        sheet.cell(row = count,column=4, value=str(score))
        sheet.cell(row = count,column=5, value=str(score))
        count += 1
        nameSet.pop(0)
    wb.save("AL人事.xlsx")
def main():#主程式啟動
    try:
        a = ['alience','checkIp','standardIp','web','chrome','score','actType']
        xmlData = readXml()
        test = xmlData.keys()
        if len(test) != len(a):
            1/0
        for i in test:
            a.remove(i)
    except:
        f = open('你他媽是不是亂改我xml內容.txt','a')
        f.write("幹給我好好看說明喔")
        f.close()
    nameSet = set() #名字set
    ops = webdriver.ChromeOptions()
    ops.binary_location = xmlData['chrome'] #chrome地址到時候用xml進行輸入
    driver = webdriver.Chrome(ops)
    #輸入網頁的部分
    driver.get(xmlData['web']) #到時候改成用xml輸入
    wait = WebDriverWait(driver, 20)
    wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'rs-table-body-wheel-area')))
    time.sleep(5)
    page_content = driver.page_source
    soup = BeautifulSoup(page_content, 'html.parser')
    data = soup.find('div', class_='rs-table-body-wheel-area')
    playerList = data.find_all('div', class_='rs-table-row') #可迭代物件 #這邊是已經爬到每位玩家的資料了
    for player in playerList:
        if xmlData['alience'] in str(player): #先用暴力法到時候速度太慢再說 #公會名稱到時候用xml輸入
            name = player.find('div', class_='rs-table-cell-group rs-table-cell-group-fixed-left')
            name = player.find('div', class_="rs-table-cell-content")
            if xmlData['checkIp']:
                if ipCheck(player,xmlData['standardIp']):
                    nameSet.add(str(name.text).upper())
            else:
                nameSet.add(str(name.text).upper())          
    #模仿人工點擊下一頁，抓取下一頁資料重複操作到沒有下一頁即可
    tab_element = driver.find_element(By.CLASS_NAME, 'rs-table-pagination-toolbar')#轉換分頁
    tab_element = tab_element.find_element(By.CLASS_NAME, 'rs-table-pagination-end')
    tab_element = tab_element.find_elements(By.CLASS_NAME, 'rs-pagination-btn')[-2]
    tab_element = tab_element.find_element(By.TAG_NAME,'a')
    tab_element.click()
    while tab_element:
        wait = WebDriverWait(driver, 20)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'rs-table-body-wheel-area')))
        page_content = driver.page_source
        soup = BeautifulSoup(page_content, 'html.parser')
        data = soup.find('div', class_='rs-table-body-wheel-area')
        playerList = data.find_all('div', class_='rs-table-row') #可迭代物件 #這邊是已經爬到每位玩家的資料了
        for player in playerList:
            if xmlData['alience'] in str(player): #先用暴力法到時候速度太慢再說 #公會名稱到時候用xml輸入
                name = player.find('div', class_='rs-table-cell-group rs-table-cell-group-fixed-left')
                name = player.find('div', class_="rs-table-cell-content")
                if xmlData['checkIp']:
                    if ipCheck(player,xmlData['standardIp']):
                        nameSet.add(str(name.text).upper())
                else:
                    nameSet.add(str(name.text).upper())    
        tab_element = driver.find_element(By.CLASS_NAME, 'rs-table-pagination-toolbar')#轉換分頁
        tab_element = tab_element.find_element(By.CLASS_NAME, 'rs-table-pagination-end')
        tab_elements = tab_element.find_elements(By.CLASS_NAME, 'rs-pagination-btn')
        try:
            tab_element.find_element(By.CLASS_NAME, 'rs-pagination-btn-disabled')
            break
        except Exception as E:
            # print(E)
            pass
        tab_element = tab_elements[-2].find_element(By.TAG_NAME,'a')
        tab_element.click()
    readExcel(nameSet,int(xmlData['score']))
    driver.quit()
    
try:
    main()
except Exception as e:
    f = open('錯誤訊息.txt','w')
    f.write(i)
    f.close()